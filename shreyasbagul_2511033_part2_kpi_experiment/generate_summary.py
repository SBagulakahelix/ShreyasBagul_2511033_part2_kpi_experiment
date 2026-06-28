import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

project_dir = r"C:\Users\bagul\OneDrive\Desktop\assignment\shreyasbagul_2511033_part2_kpi_experiment"
clean_path = os.path.join(project_dir, "analysis", "experiment_analysis.xlsx")
summary_path = os.path.join(project_dir, "outputs", "experiment_summary.xlsx")

# Load prepared data
df = pd.read_excel(clean_path, sheet_name="experiment_data")

print("Generating Experiment Summary Workbook...")

# Create Workbook
wb = openpyxl.Workbook()
# remove default sheet
default_sheet = wb.active
wb.remove(default_sheet)

# Style Definitions
font_family = "Segoe UI"
title_font = Font(name=font_family, size=14, bold=True, color="FFFFFF")
subtitle_font = Font(name=font_family, size=10, italic=True, color="FFFFFF")
header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
data_font = Font(name=font_family, size=10)
bold_font = Font(name=font_family, size=10, bold=True)
total_font = Font(name=font_family, size=10, bold=True, color="1F4E79")

# Colors
title_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid") # Deep steel blue
header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid") # Medium steel blue
zebra_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid") # Light grey
total_fill = PatternFill(start_color="E9ECEF", end_color="E9ECEF", fill_type="solid") # Darker grey for totals

# Borders
thin_border = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)

# Helper function to format layout
def format_sheet_layout(ws, freeze_row=4):
    ws.views.sheetView[0].showGridLines = True
    if freeze_row > 0:
        ws.freeze_panes = f"A{freeze_row}"
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = str(cell.value or '')
            if cell.number_format and '%' in cell.number_format:
                val += '%'
            if '\n' in val:
                lines = val.split('\n')
                max_len = max(max_len, max(len(l) for l in lines))
            else:
                max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

# Helper function to write title
def write_title_block(ws, title, subtitle, max_cols=5):
    col_letter_end = get_column_letter(max_cols)
    ws.merge_cells(f"A1:{col_letter_end}1")
    ws.merge_cells(f"A2:{col_letter_end}2")
    ws["A1"] = title
    ws["A1"].font = title_font
    ws["A1"].fill = title_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A2"] = subtitle
    ws["A2"].font = subtitle_font
    ws["A2"].fill = title_fill
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 18

# Helper to write headers
def write_headers(ws, headers, start_row=4):
    ws.row_dimensions[start_row].height = 22
    for c_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=c_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Helper to style data row
def style_row(ws, row_idx, num_cols, align_left_cols=[1], number_formats={}):
    ws.row_dimensions[row_idx].height = 20
    is_zebra = (row_idx % 2 == 0)
    for c_idx in range(1, num_cols + 1):
        cell = ws.cell(row=row_idx, column=c_idx)
        cell.font = data_font
        cell.border = thin_border
        if is_zebra:
            cell.fill = zebra_fill
            
        # Alignment
        if c_idx in align_left_cols:
            cell.alignment = Alignment(horizontal="left", vertical="center")
        else:
            cell.alignment = Alignment(horizontal="right", vertical="center")
            
        # Formatting
        if c_idx in number_formats:
            cell.number_format = number_formats[c_idx]

# Total Border Definition
total_border = Border(
    top=Side(style='thin', color='1F4E79'),
    bottom=Side(style='double', color='1F4E79'),
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9')
)

# Helper function to write total row
def write_total_row(ws, row_idx, num_cols, label_col=1, formula_dict={}):
    ws.row_dimensions[row_idx].height = 22
    for c_idx in range(1, num_cols + 1):
        cell = ws.cell(row=row_idx, column=c_idx)
        cell.font = total_font
        cell.fill = total_fill
        cell.border = total_border
        
        if c_idx == label_col:
            cell.value = "Grand Total"
            cell.alignment = Alignment(horizontal="left", vertical="center")
        elif c_idx in formula_dict:
            cell.value = formula_dict[c_idx]
            cell.alignment = Alignment(horizontal="right", vertical="center")
        else:
            cell.value = ""

# ==========================================
# SHEET 1: Overall_Comparison
# ==========================================
print("  - Overall_Comparison")
ws_overall = wb.create_sheet("Overall_Comparison")
write_title_block(ws_overall, "OVERALL EXPERIMENT METRICS COMPARISON", "Control vs Treatment (Deduplicated Users N=1400)", max_cols=5)

headers_overall = ["Experiment Metric", "Control Group", "Treatment Group", "Absolute Difference", "Percentage Lift"]
write_headers(ws_overall, headers_overall)

# We will write the calculated values for Control and Treatment, and write formulas for Diff and Lift
# To make it easy, we calculate them in python and write them
control_df = df[df['experiment_group'] == 'Control']
treatment_df = df[df['experiment_group'] == 'Treatment']

metrics_raw = [
    ("User Count", len(control_df), len(treatment_df), "#,##0"),
    ("Landing Page Visit Rate", control_df['visited_landing_page'].mean(), treatment_df['visited_landing_page'].mean(), "0.00%"),
    ("Trial Start Rate", control_df['started_trial'].mean(), treatment_df['started_trial'].mean(), "0.00%"),
    ("Onboarding Completion Rate", control_df['completed_onboarding'].mean(), treatment_df['completed_onboarding'].mean(), "0.00%"),
    ("Paid Conversion Rate", control_df['converted_to_paid'].mean(), treatment_df['converted_to_paid'].mean(), "0.00%"),
    ("Average Revenue Per User (ARPU)", control_df['revenue_30d'].mean(), treatment_df['revenue_30d'].mean(), "$#,##0.00"),
    ("Average Revenue Per Converted User", control_df[control_df['converted_to_paid'] == 1]['revenue_30d'].mean(), treatment_df[treatment_df['converted_to_paid'] == 1]['revenue_30d'].mean(), "$#,##0.00"),
    ("Refund Rate", control_df['refund_requested'].mean(), treatment_df['refund_requested'].mean(), "0.00%"),
    ("Support Ticket Rate", (control_df['support_tickets_30d'] > 0).mean(), (treatment_df['support_tickets_30d'] > 0).mean(), "0.00%"),
    ("Average Engagement Score", control_df['engagement_score'].mean(), treatment_df['engagement_score'].mean(), "0.00"),
    ("Average Days to Convert", control_df[control_df['converted_to_paid'] == 1]['days_to_convert'].mean(), treatment_df[treatment_df['converted_to_paid'] == 1]['days_to_convert'].mean(), "0.0")
]

for idx, (metric_name, ctrl_val, treat_val, num_fmt) in enumerate(metrics_raw, 5):
    ws_overall.cell(row=idx, column=1, value=metric_name)
    ws_overall.cell(row=idx, column=2, value=ctrl_val)
    ws_overall.cell(row=idx, column=3, value=treat_val)
    
    # Formulas for Diff and Lift
    # For User Count: Diff is C-B, Lift is (C-B)/B
    # For average revenue/days: Diff is C-B
    # For rates, Diff is C-B, but let's represent it nicely
    ws_overall.cell(row=idx, column=4, value=f"=C{idx}-B{idx}")
    
    if metric_name == "User Count":
        ws_overall.cell(row=idx, column=5, value=f"=(C{idx}-B{idx})/B{idx}")
    elif "Rate" in metric_name or "ARPU" in metric_name or "Revenue" in metric_name or "Engagement" in metric_name or "Days" in metric_name:
        ws_overall.cell(row=idx, column=5, value=f"=(C{idx}-B{idx})/B{idx}")
    else:
        ws_overall.cell(row=idx, column=5, value="")
        
    style_row(ws_overall, idx, 5, align_left_cols=[1], number_formats={2: num_fmt, 3: num_fmt, 4: num_fmt, 5: "0.00%"})
    
    # For Absolute Difference of percentages, show as percentage point diff (e.g. +3.85%)
    # Let's override the number format for Diff if it is a rate
    if "Rate" in metric_name:
        # Show absolute difference in percentage points (using the same 0.00% format)
        ws_overall.cell(row=idx, column=4).number_format = "+0.00%;-0.00%;0.00%"
    elif "ARPU" in metric_name or "Revenue" in metric_name:
        ws_overall.cell(row=idx, column=4).number_format = "+$#,##0.00;-$#,##0.00;$0.00"
    elif "Score" in metric_name or "Days" in metric_name:
        ws_overall.cell(row=idx, column=4).number_format = "+0.0;-0.0;0.0"
        
    # Lift formatting
    ws_overall.cell(row=idx, column=5).number_format = "+0.00%;-0.00%;0.00%"

format_sheet_layout(ws_overall, freeze_row=5)

# ==========================================
# Helper to write segment sheets
# ==========================================
def write_segment_sheet(sheet_name, col_name, title):
    print(f"  - {sheet_name}")
    ws = wb.create_sheet(sheet_name)
    write_title_block(ws, title, "Experiment Segment Breakdown - Shreyas Bagul (2511033)", max_cols=8)
    
    headers_seg = [
        col_name.replace('_', ' ').title(), 
        "Control N", "Treatment N", 
        "Control CR", "Treatment CR", "CR Lift",
        "Control ARPU", "Treatment ARPU", "ARPU Lift"
    ]
    write_headers(ws, headers_seg)
    
    # Group by segment
    segments = sorted(df[col_name].dropna().unique())
    
    for idx, seg in enumerate(segments, 5):
        df_seg = df[df[col_name] == seg]
        control_seg = df_seg[df_seg['experiment_group'] == 'Control']
        treatment_seg = df_seg[df_seg['experiment_group'] == 'Treatment']
        
        n_c = len(control_seg)
        n_t = len(treatment_seg)
        cr_c = control_seg['converted_to_paid'].mean() if n_c > 0 else 0
        cr_t = treatment_seg['converted_to_paid'].mean() if n_t > 0 else 0
        arpu_c = control_seg['revenue_30d'].mean() if n_c > 0 else 0
        arpu_t = treatment_seg['revenue_30d'].mean() if n_t > 0 else 0
        
        ws.cell(row=idx, column=1, value=seg)
        ws.cell(row=idx, column=2, value=n_c)
        ws.cell(row=idx, column=3, value=n_t)
        ws.cell(row=idx, column=4, value=cr_c)
        ws.cell(row=idx, column=5, value=cr_t)
        
        # Formulas for lifts
        # CR Lift: (E - D) / D
        ws.cell(row=idx, column=6, value=f"=IF(D{idx}=0, \"\", (E{idx}-D{idx})/D{idx})")
        
        ws.cell(row=idx, column=7, value=arpu_c)
        ws.cell(row=idx, column=8, value=arpu_t)
        
        # ARPU Lift: (H - G) / G
        ws.cell(row=idx, column=9, value=f"=IF(G{idx}=0, \"\", (H{idx}-G{idx})/G{idx})")
        
        style_row(
            ws, idx, 9, 
            align_left_cols=[1], 
            number_formats={
                2: "#,##0", 3: "#,##0", 
                4: "0.00%", 5: "0.00%", 6: "+0.00%;-0.00%;0.00%",
                7: "$#,##0.00", 8: "$#,##0.00", 9: "+0.00%;-0.00%;0.00%"
            }
        )
        
    # Total row at bottom
    total_row_idx = 5 + len(segments)
    
    # Formulas for totals
    write_total_row(ws, total_row_idx, 9, label_col=1, formula_dict={
        2: f"=SUM(B5:B{total_row_idx-1})",
        3: f"=SUM(C5:C{total_row_idx-1})",
        # Weighted average CR: sum of conversions / total N
        # We can't do simple SUM, so let's calculate directly or write direct formulas
        # Actually, let's write a formula that computes the overall rates for the group
        # Wait, since N is in B and CR is in D, total conversions is SUMPRODUCT(B, D)
        # So overall CR is SUMPRODUCT(B, D) / SUM(B)
        4: f"=SUMPRODUCT(B5:B{total_row_idx-1}, D5:D{total_row_idx-1})/B{total_row_idx}",
        5: f"=SUMPRODUCT(C5:C{total_row_idx-1}, E5:E{total_row_idx-1})/C{total_row_idx}",
        6: f"=(E{total_row_idx}-D{total_row_idx})/D{total_row_idx}",
        # Weighted average ARPU: SUMPRODUCT(B, G) / SUM(B)
        7: f"=SUMPRODUCT(B5:B{total_row_idx-1}, G5:G{total_row_idx-1})/B{total_row_idx}",
        8: f"=SUMPRODUCT(C5:C{total_row_idx-1}, H5:H{total_row_idx-1})/C{total_row_idx}",
        9: f"=(H{total_row_idx}-G{total_row_idx})/G{total_row_idx}"
    })
    
    # Format total row
    ws.cell(row=total_row_idx, column=2).number_format = "#,##0"
    ws.cell(row=total_row_idx, column=3).number_format = "#,##0"
    ws.cell(row=total_row_idx, column=4).number_format = "0.00%"
    ws.cell(row=total_row_idx, column=5).number_format = "0.00%"
    ws.cell(row=total_row_idx, column=6).number_format = "+0.00%;-0.00%;0.00%"
    ws.cell(row=total_row_idx, column=7).number_format = "$#,##0.00"
    ws.cell(row=total_row_idx, column=8).number_format = "$#,##0.00"
    ws.cell(row=total_row_idx, column=9).number_format = "+0.00%;-0.00%;0.00%"
    
    format_sheet_layout(ws, freeze_row=5)

# Write all segment sheets
write_segment_sheet("Region_Analysis", "region", "EXPERIMENT SEGMENT ANALYSIS: REGION")
write_segment_sheet("Device_Analysis", "device_type", "EXPERIMENT SEGMENT ANALYSIS: DEVICE TYPE")
write_segment_sheet("Traffic_Source_Analysis", "traffic_source", "EXPERIMENT SEGMENT ANALYSIS: TRAFFIC SOURCE")
write_segment_sheet("Plan_Type_Analysis", "plan_type", "EXPERIMENT SEGMENT ANALYSIS: PLAN TYPE")

# Save Workbook
print(f"Saving experiment summary workbook to {summary_path}...")
wb.save(summary_path)
print("Experiment Summary Workbook generated successfully!")
